from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_BASE_URL = os.getenv("TEABLE_BASE_URL", "https://app.teable.ai")
DEFAULT_TOKEN = os.getenv("TEABLE_TOKEN", "")
DEFAULT_TABLE_ID = os.getenv("TEABLE_SALES_TABLE_ID", os.getenv("TEABLE_TABLE_ID", ""))
DEFAULT_VIEW_ID = os.getenv("TEABLE_SALES_VIEW_ID", os.getenv("TEABLE_VIEW_ID", ""))
DEFAULT_OUTPUT = os.getenv("SALES_REPORT_OUTPUT", "業績明細圖表.xlsx")

FIELD_MAP = {
    "date": os.getenv("FIELD_DATE", "日期"),
    "customer": os.getenv("FIELD_CUSTOMER", "客戶"),
    "factory": os.getenv("FIELD_FACTORY", "工廠"),
    "sales": os.getenv("FIELD_SALES", "業務"),
    "amount": os.getenv("FIELD_AMOUNT", "出貨金額"),
    "hold": os.getenv("FIELD_HOLD", "HOLD金額"),
    "discount": os.getenv("FIELD_DISCOUNT", "折讓"),
    "net": os.getenv("FIELD_NET", "淨出貨金額"),
    "status": os.getenv("FIELD_STATUS", "狀態"),
    "order_no": os.getenv("FIELD_ORDER_NO", "訂單號"),
    "part_no": os.getenv("FIELD_PART_NO", "料號"),
    "remark": os.getenv("FIELD_REMARK", "備註"),
}


def _safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("NT$", "").replace("$", "")
    if text in {"", "-", "None", "nan"}:
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join([_to_text(v) for v in value if v is not None])
    return str(value).strip()


def _to_date(value: Any) -> Optional[pd.Timestamp]:
    if value is None or value == "":
        return None
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return None
        return ts
    except Exception:
        return None


def fetch_teable_records(
    token: str,
    table_id: str,
    base_url: str = DEFAULT_BASE_URL,
    view_id: str = "",
    page_size: int = 1000,
) -> List[Dict[str, Any]]:
    if not token:
        raise ValueError("缺少 TEABLE_TOKEN")
    if not table_id:
        raise ValueError("缺少 TEABLE_SALES_TABLE_ID 或 TEABLE_TABLE_ID")

    base_url = base_url.rstrip("/")
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    offset = 0
    out: List[Dict[str, Any]] = []

    while True:
        params: Dict[str, Any] = {"take": page_size, "skip": offset}
        if view_id:
            params["viewId"] = view_id

        url = f"{base_url}/api/table/{table_id}/record"
        resp = requests.get(url, headers=headers, params=params, timeout=60)
        resp.raise_for_status()
        data = resp.json()

        records = data.get("records") or data.get("data", {}).get("records") or []
        if not records:
            break

        out.extend(records)
        if len(records) < page_size:
            break
        offset += page_size

    return out


def teable_records_to_df(records: List[Dict[str, Any]]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for rec in records:
        fields = rec.get("fields", {}) if isinstance(rec, dict) else {}
        row = {"_record_id": rec.get("id", "")}
        row.update(fields)
        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(
            columns=[
                FIELD_MAP["date"],
                FIELD_MAP["customer"],
                FIELD_MAP["factory"],
                FIELD_MAP["sales"],
                FIELD_MAP["amount"],
                FIELD_MAP["hold"],
                FIELD_MAP["discount"],
                FIELD_MAP["net"],
            ]
        )
    return df


def normalize_sales_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    for col in FIELD_MAP.values():
        if col not in out.columns:
            out[col] = ""

    out["_date"] = out[FIELD_MAP["date"]].apply(_to_date)
    out["_month"] = out["_date"].apply(lambda x: x.strftime("%Y-%m") if x is not None else "")
    out["_customer"] = out[FIELD_MAP["customer"]].apply(_to_text)
    out["_factory"] = out[FIELD_MAP["factory"]].apply(_to_text)
    out["_sales"] = out[FIELD_MAP["sales"]].apply(_to_text)
    out["_status"] = out[FIELD_MAP["status"]].apply(_to_text)

    out["_amount"] = out[FIELD_MAP["amount"]].apply(_safe_float)
    out["_hold"] = out[FIELD_MAP["hold"]].apply(_safe_float)
    out["_discount"] = out[FIELD_MAP["discount"]].apply(_safe_float)

    raw_net = out[FIELD_MAP["net"]].apply(_safe_float)
    out["_net"] = raw_net.where(raw_net != 0, out["_amount"] - out["_hold"] - out["_discount"])
    out["_date_str"] = out["_date"].apply(lambda x: x.strftime("%Y-%m-%d") if x is not None else "")

    return out


def filter_month(df: pd.DataFrame, month_str: str) -> pd.DataFrame:
    if not month_str:
        return df.copy()
    return df[df["_month"] == month_str].copy()


def summarize_sales(df: pd.DataFrame) -> Dict[str, Any]:
    total_amount = float(df["_amount"].sum()) if not df.empty else 0.0
    total_hold = float(df["_hold"].sum()) if not df.empty else 0.0
    total_discount = float(df["_discount"].sum()) if not df.empty else 0.0
    total_net = float(df["_net"].sum()) if not df.empty else 0.0

    by_customer = (
        df.groupby("_customer", dropna=False)["_net"]
        .sum()
        .reset_index()
        .rename(columns={"_customer": "客戶", "_net": "淨出貨金額"})
        .sort_values("淨出貨金額", ascending=False)
    )

    by_factory = (
        df.groupby("_factory", dropna=False)["_net"]
        .sum()
        .reset_index()
        .rename(columns={"_factory": "工廠", "_net": "淨出貨金額"})
        .sort_values("淨出貨金額", ascending=False)
    )

    by_sales = (
        df.groupby("_sales", dropna=False)["_net"]
        .sum()
        .reset_index()
        .rename(columns={"_sales": "業務", "_net": "淨出貨金額"})
        .sort_values("淨出貨金額", ascending=False)
    )

    daily = (
        df.groupby("_date_str", dropna=False)[["_amount", "_hold", "_discount", "_net"]]
        .sum()
        .reset_index()
        .rename(
            columns={
                "_date_str": "日期",
                "_amount": "出貨金額",
                "_hold": "HOLD金額",
                "_discount": "折讓",
                "_net": "淨出貨金額",
            }
        )
        .sort_values("日期")
    )

    return {
        "total_amount": total_amount,
        "total_hold": total_hold,
        "total_discount": total_discount,
        "total_net": total_net,
        "customer_count": int(df["_customer"].replace("", pd.NA).dropna().nunique()),
        "factory_count": int(df["_factory"].replace("", pd.NA).dropna().nunique()),
        "sales_count": int(df["_sales"].replace("", pd.NA).dropna().nunique()),
        "by_customer": by_customer,
        "by_factory": by_factory,
        "by_sales": by_sales,
        "daily": daily,
    }


def _fmt_money(value: float, currency_symbol: str = "NT$") -> str:
    return f"{currency_symbol} {value:,.0f}"


def build_excel_report(
    raw_df: pd.DataFrame,
    month_df: pd.DataFrame,
    summary: Dict[str, Any],
    report_month: str,
    company_name: str = "",
    currency_symbol: str = "NT$",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "業績圖表總覽"

    header_fill = PatternFill("solid", fgColor="1F2937")
    white_font = Font(color="FFFFFF", bold=True, size=16)
    title_font = Font(bold=True, size=18)

    ws["A1"] = f"{report_month} 業績報表"
    ws["A1"].font = title_font
    ws["A2"] = f"子表 {company_name}" if company_name else "子表"
    ws["A4"] = "總業績"
    ws["C4"] = "客戶數"
    ws["E4"] = "廠商數"
    ws["G4"] = "淨出貨"

    for cell in ["A4", "C4", "E4", "G4"]:
        ws[cell].fill = header_fill
        ws[cell].font = white_font
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    ws["A5"] = summary["total_amount"]
    ws["C5"] = summary["customer_count"]
    ws["E5"] = summary["factory_count"]
    ws["G5"] = summary["total_net"]

    ws["A5"].number_format = f'"{currency_symbol}" #,##0'
    ws["G5"].number_format = f'"{currency_symbol}" #,##0'

    by_customer = summary["by_customer"].copy()
    if by_customer.empty:
        by_customer = pd.DataFrame({"客戶": ["N/A"], "淨出貨金額": [0.0]})

    ws["A8"] = "客戶業績比較"
    ws["F8"] = "業績佔比"
    ws["A8"].font = Font(bold=True, size=12)
    ws["F8"].font = Font(bold=True, size=12)

    ws["A9"] = "客戶"
    ws["B9"] = "淨出貨金額"
    max_customer_rows = min(10, len(by_customer))
    for i, (_, row) in enumerate(by_customer.head(10).iterrows(), start=10):
        ws[f"A{i}"] = row["客戶"] or "(空白)"
        ws[f"B{i}"] = float(row["淨出貨金額"])
        ws[f"B{i}"].number_format = f'"{currency_symbol}" #,##0'

    end_customer_row = 9 + max(1, max_customer_rows)
    data = Reference(ws, min_col=2, min_row=9, max_row=end_customer_row)
    cats = Reference(ws, min_col=1, min_row=10, max_row=end_customer_row)

    bar = BarChart()
    bar.title = "客戶業績比較"
    bar.y_axis.title = "金額"
    bar.height = 8
    bar.width = 13
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.legend = None
    ws.add_chart(bar, "A20")

    donut = DoughnutChart()
    donut.title = "業績佔比"
    donut.holeSize = 55
    donut.height = 8
    donut.width = 11
    donut.add_data(data, titles_from_data=True)
    donut.set_categories(cats)
    ws.add_chart(donut, "F20")

    daily = summary["daily"].copy()
    if daily.empty:
        daily = pd.DataFrame({"日期": ["N/A"], "淨出貨金額": [0.0], "出貨金額": [0.0], "HOLD金額": [0.0], "折讓": [0.0]})

    start_row = 38
    ws[f"A{start_row}"] = "每日趨勢"
    ws[f"A{start_row}"].font = Font(bold=True, size=12)
    ws[f"A{start_row+1}"] = "日期"
    ws[f"B{start_row+1}"] = "淨出貨金額"
    for i, (_, row) in enumerate(daily.iterrows(), start=start_row + 2):
        ws[f"A{i}"] = row["日期"]
        ws[f"B{i}"] = float(row["淨出貨金額"])
        ws[f"B{i}"].number_format = f'"{currency_symbol}" #,##0'

    line = LineChart()
    line.title = "每日淨出貨趨勢"
    line.y_axis.title = "金額"
    line.height = 8
    line.width = 18
    end_daily_row = start_row + 1 + max(1, len(daily))
    d2 = Reference(ws, min_col=2, min_row=start_row + 1, max_row=end_daily_row)
    c2 = Reference(ws, min_col=1, min_row=start_row + 2, max_row=end_daily_row)
    line.add_data(d2, titles_from_data=True)
    line.set_categories(c2)
    ws.add_chart(line, "D38")

    detail = wb.create_sheet("業績明細表")
    keep_cols = [
        FIELD_MAP["date"], FIELD_MAP["customer"], FIELD_MAP["factory"], FIELD_MAP["sales"],
        FIELD_MAP["amount"], FIELD_MAP["hold"], FIELD_MAP["discount"], FIELD_MAP["net"],
        FIELD_MAP["status"], FIELD_MAP["order_no"], FIELD_MAP["part_no"], FIELD_MAP["remark"],
    ]
    for idx, col in enumerate(keep_cols, start=1):
        detail.cell(row=1, column=idx, value=col).font = Font(bold=True)

    for r, (_, row) in enumerate(month_df.iterrows(), start=2):
        for c, col in enumerate(keep_cols, start=1):
            value = row.get(col, "")
            detail.cell(row=r, column=c, value="" if pd.isna(value) else value)

    for col_idx in range(1, len(keep_cols) + 1):
        detail.column_dimensions[get_column_letter(col_idx)].width = 16

    def _write_df_sheet(name: str, df: pd.DataFrame):
        sh = wb.create_sheet(name)
        if df.empty:
            sh["A1"] = "無資料"
            return
        for c, col in enumerate(df.columns, start=1):
            sh.cell(row=1, column=c, value=col).font = Font(bold=True)
        for r, (_, row) in enumerate(df.iterrows(), start=2):
            for c, col in enumerate(df.columns, start=1):
                sh.cell(row=r, column=c, value="" if pd.isna(row[col]) else row[col])
        for col_idx in range(1, len(df.columns) + 1):
            sh.column_dimensions[get_column_letter(col_idx)].width = 18

    _write_df_sheet("客戶彙總", summary["by_customer"])
    _write_df_sheet("工廠彙總", summary["by_factory"])
    _write_df_sheet("業務彙總", summary["by_sales"])
    _write_df_sheet("每日明細", summary["daily"])
    _write_df_sheet("RawData", raw_df)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def render_sales_report_page():
    st.subheader("📊 業績明細表")
    st.caption("保留原 Excel 明細架構，同時新增圖表首頁。資料來源為 Teable。")

    with st.expander("設定", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            token = st.text_input("TEABLE_TOKEN", value=DEFAULT_TOKEN, type="password")
            table_id = st.text_input("Teable Sales Table ID", value=DEFAULT_TABLE_ID)
            view_id = st.text_input("Teable Sales View ID", value=DEFAULT_VIEW_ID)
        with c2:
            base_url = st.text_input("Teable Base URL", value=DEFAULT_BASE_URL)
            now = datetime.now()
            month_default = os.getenv("REPORT_MONTH", now.strftime("%Y-%m"))
            report_month = st.text_input("報表月份 (YYYY-MM)", value=month_default)
            company_name = st.text_input("子表名稱 / 公司名稱", value=os.getenv("COMPANY_NAME", ""))
            currency_symbol = st.text_input("幣別符號", value=os.getenv("CURRENCY_SYMBOL", "NT$"))

    if not st.button("讀取並產生業績報表", type="primary", use_container_width=True):
        st.info("按上方按鈕後，會從 Teable 抓資料並顯示圖表，同時可下載 Excel。")
        return

    try:
        with st.spinner("從 Teable 抓資料中..."):
            records = fetch_teable_records(token=token, table_id=table_id, base_url=base_url, view_id=view_id)
            raw_df = teable_records_to_df(records)
            norm_df = normalize_sales_df(raw_df)
            month_df = filter_month(norm_df, report_month)
            summary = summarize_sales(month_df)

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("總業績", _fmt_money(summary["total_amount"], currency_symbol))
        k2.metric("客戶數", summary["customer_count"])
        k3.metric("廠商數", summary["factory_count"])
        k4.metric("淨出貨", _fmt_money(summary["total_net"], currency_symbol))

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**客戶業績比較**")
            if summary["by_customer"].empty:
                st.write("無資料")
            else:
                st.bar_chart(summary["by_customer"].head(10).set_index("客戶"))

        with c2:
            st.markdown("**業績佔比**")
            if summary["by_customer"].empty:
                st.write("無資料")
            else:
                pie_df = summary["by_customer"].head(10).copy()
                total = max(pie_df["淨出貨金額"].sum(), 1)
                pie_df["佔比%"] = pie_df["淨出貨金額"] / total * 100
                st.dataframe(
                    pie_df.assign(淨出貨金額=pie_df["淨出貨金額"].map(lambda x: _fmt_money(x, currency_symbol))),
                    use_container_width=True,
                    hide_index=True,
                )

        st.markdown("**每日趨勢**")
        if summary["daily"].empty:
            st.write("無資料")
        else:
            st.line_chart(summary["daily"].set_index("日期")[["淨出貨金額"]])

        t1, t2, t3 = st.tabs(["業績明細表", "客戶/工廠/業務彙總", "RawData"])
        with t1:
            keep_cols = [
                FIELD_MAP["date"], FIELD_MAP["customer"], FIELD_MAP["factory"], FIELD_MAP["sales"],
                FIELD_MAP["amount"], FIELD_MAP["hold"], FIELD_MAP["discount"], FIELD_MAP["net"],
                FIELD_MAP["status"], FIELD_MAP["order_no"], FIELD_MAP["part_no"], FIELD_MAP["remark"],
            ]
            show_df = month_df[keep_cols].copy() if not month_df.empty else pd.DataFrame(columns=keep_cols)
            st.dataframe(show_df, use_container_width=True, hide_index=True)

        with t2:
            a, b, c = st.columns(3)
            with a:
                st.markdown("**客戶彙總**")
                st.dataframe(summary["by_customer"], use_container_width=True, hide_index=True)
            with b:
                st.markdown("**工廠彙總**")
                st.dataframe(summary["by_factory"], use_container_width=True, hide_index=True)
            with c:
                st.markdown("**業務彙總**")
                st.dataframe(summary["by_sales"], use_container_width=True, hide_index=True)

        with t3:
            st.dataframe(raw_df, use_container_width=True, hide_index=True)

        excel_bytes = build_excel_report(
            raw_df=raw_df,
            month_df=month_df,
            summary=summary,
            report_month=report_month,
            company_name=company_name,
            currency_symbol=currency_symbol,
        )
        st.download_button(
            "下載 Excel 業績圖表報表",
            data=excel_bytes,
            file_name=DEFAULT_OUTPUT,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception as e:
        st.error(f"產生業績報表失敗：{e}")
