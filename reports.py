# -*- coding: utf-8 -*-
"""
reports.py  ── GLOCOM Control Tower 業績明細表
v4：加入統計圖表
  - 已出貨 vs 預計出貨（水平堆疊長條）
  - 依工廠 / 依客戶銷貨佔比（甜甜圈圓餅）
  - 近12個月銷貨趨勢（長條 + 折線）

[修正 v4.1]
  - build_subset_mask "unshipped" 模式移除 year_ok 年份篩選

[修正 v4.2]
  - 業績明細表月份選單預設改為當月

[修正 v4.3]
  - SANDY_NEW_ORDER_SPECS 加入 Working Gerber Approval、Engineering Question
  - SANDY_INTERNAL_WIP_SPECS 加入 Working Gerber Approval、Engineering Question
  - render_teable_subset_table 加入 Excel 下載（自動欄寬）

[修正 v5.0]
  - render_sales_detail_from_teable 加入三個 Tab：
      業績明細 / 報關單價 / 匯 HK OBU 金額
"""

from __future__ import annotations

import io
import re
import json

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# ── 報關單價 & OBU 模組（新增）──────────────────────────────────────────────
try:
    from obu_page import render_customs_price_tab, render_obu_calc_tab
    _OBU_AVAILABLE = True
except ImportError:
    _OBU_AVAILABLE = False

# ================================
# 歷史移轉金額（一次性，寫死）
# key: "YYYY-MM"  value: float（已出貨）
# ================================
LEGACY_SHIPPED: dict[str, float] = {
    "2026-03": 58_392.42,
}

# ================================
# FIELD CANDIDATES
# ================================
PO_CANDIDATES        = ["PO#", "PO", "P/O", "訂單編號", "訂單號", "訂單號碼", "工單", "工單號", "單號"]
CUSTOMER_CANDIDATES  = ["Customer", "客戶", "客戶名稱"]
PART_CANDIDATES      = ["Part No", "Part No.", "P/N", "客戶料號", "Cust. P / N", "LS P/N",
                        "料號", "品號", "成品料號", "產品料號"]
QTY_CANDIDATES       = ["Qty", "Order Q'TY (PCS)", "Order Q'TY\n (PCS)", "訂購量 (PCS)",
                        "訂購量", "Q'TY", "數量", "PCS", "訂單量", "生產數量", "投產數"]
FACTORY_CANDIDATES   = ["Factory", "工廠", "廠編"]
WIP_CANDIDATES       = ["WIP", "WIP Stage", "進度", "製程", "工序", "目前站別", "生產進度"]
FACTORY_DUE_CANDIDATES = ["Factory Due Date", "工廠交期", "交貨日期", "Required Ship date",
                           "confrimed DD", "交期", "預交日", "預定交期", "交貨期"]
REMARK_CANDIDATES      = ["Remark", "備註", "情況", "備註說明", "Note", "說明", "異常備註"]
ORDER_DATE_CANDIDATES  = ["客戶下單日期", "工廠下單日期", "下單日期", "Order Date",
                           "PO Date", "Date", "訂單日期", "接單日期"]
AMOUNT_ORDER_CANDIDATES = ["接單金額", "接單總金額", "Order Amount", "Order amount",
                            "Order Total", "客戶金額", "Sales Amount", "Quote Total",
                            "Total Amount", "Amount"]
AMOUNT_SHIP_CANDIDATES  = ["銷貨金額", "出貨金額", "出貨總金額", "Shipment Amount",
                            "Ship Amount", "Shipping Amount", "Invoice Amount",
                            "Invoice Total", "出貨發票金額", "Invoice", "INVOICE"]
ACTUAL_SHIP_DATE_CANDIDATES  = ["出貨日期_排序", "出貨日期", "Actual Ship Date", "Actual ship date"]
PLANNED_SHIP_DATE_CANDIDATES = ["Ship date", "Ship Date", "Required Ship date", "confrimed DD"]

CANCELLED_KEYWORDS = ["PO CANCELLED", "PO CANCELED", "CANCELLATION",
                      "CANCELLED", "CANCELED", "CANCEL"]

# 圖表色盤
CHART_COLORS = ["#378ADD", "#1D9E75", "#D85A30", "#D4537E",
                "#7F77DD", "#888780", "#639922", "#BA7517"]


# ================================
# UTILITIES
# ================================

def _norm(text: str) -> str:
    s = str(text or "")
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def make_unique_columns(columns):
    seen, out = {}, []
    for c in list(columns):
        name = str(c)
        n = seen.get(name, 0)
        out.append(name if n == 0 else f"{name}_{n + 1}")
        seen[name] = n + 1
    return out


def col_candidates(*names):
    return [str(x).strip() for x in names if str(x).strip()]


def get_series_by_col(df: pd.DataFrame, col_name: str | None):
    if df is None or df.empty or not col_name or col_name not in df.columns:
        return None
    obj = df[col_name]
    return obj.iloc[:, 0] if isinstance(obj, pd.DataFrame) else obj


def find_col(df: pd.DataFrame, candidates):
    if df is None or df.empty:
        return None
    norm_map = {_norm(c): c for c in df.columns}
    for cand in candidates:
        if _norm(cand) in norm_map:
            return norm_map[_norm(cand)]
    for cand in candidates:
        n = _norm(cand)
        for col in df.columns:
            if n and n in _norm(col):
                return col
    return None


def parse_amount_series(series: pd.Series | None) -> pd.Series:
    if series is None:
        return pd.Series(dtype="float64")
    s = series.astype(str).fillna("").str.strip()
    s = s.str.replace(r"[^0-9.\-]", "", regex=True)
    s = s.replace({"": None, ".": None, "-": None})
    return pd.to_numeric(s, errors="coerce")


def _to_naive_ns(parsed: pd.Series) -> "np.ndarray":
    """把 datetime Series 轉成 tz-naive datetime64[ns] numpy array，相容新版 pandas。"""
    if parsed.empty:
        return parsed.values.astype("datetime64[ns]")
    try:
        if getattr(parsed.dt, "tz", None) is not None:
            parsed = parsed.dt.tz_convert("UTC").dt.tz_localize(None)
    except Exception:
        pass
    return parsed.values.astype("datetime64[ns]")


def parse_mixed_date_series(series: pd.Series | None) -> pd.Series:
    """
    混合格式日期解析，相容 Python 3.14 + 新版 pandas。
    """
    import numpy as np

    if series is None:
        return pd.Series(dtype="datetime64[ns]")

    s = series.astype(str).fillna("").str.strip()
    s = s.replace({"": None, "nan": None, "NaT": None, "None": None})

    result = np.full(len(s), np.datetime64("NaT"), dtype="datetime64[ns]")
    s_vals  = s.values
    s_notna = s.notna().values

    # ① Excel serial number
    nums = pd.to_numeric(s, errors="coerce")
    mask_num = (nums.notna() & (nums > 20000) & (nums < 80000)).values
    if mask_num.any():
        parsed = pd.to_datetime(nums[mask_num], unit="D", origin="1899-12-30", errors="coerce")
        result[mask_num] = _to_naive_ns(parsed)

    # ② 固定格式
    for fmt in ["%Y-%m-%d", "%y-%m-%d"]:
        rem = np.isnat(result) & s_notna
        if rem.any():
            parsed = pd.to_datetime(pd.Series(s_vals[rem]), format=fmt, errors="coerce")
            result[rem] = _to_naive_ns(parsed)

    # ③ 英文月份
    rem = np.isnat(result) & s_notna
    if rem.any():
        cleaned = (pd.Series(s_vals[rem])
                   .str.replace(".", "", regex=False)
                   .str.replace(",", "", regex=False)
                   .str.replace(r"\s+", " ", regex=True).str.strip())
        parsed = pd.to_datetime(cleaned, format="%b %d %y", errors="coerce")
        result[rem] = _to_naive_ns(parsed)

    # ④ fallback
    rem = np.isnat(result) & s_notna
    if rem.any():
        parsed = pd.to_datetime(pd.Series(s_vals[rem]), errors="coerce")
        result[rem] = _to_naive_ns(parsed)

    return pd.Series(result, index=series.index, dtype="datetime64[ns]")


def _is_cancelled(wip_series: pd.Series) -> pd.Series:
    upper = wip_series.astype(str).str.strip().str.upper()
    mask  = pd.Series(False, index=wip_series.index)
    for kw in CANCELLED_KEYWORDS:
        mask = mask | upper.str.contains(kw.upper(), regex=False, na=False)
    return mask


# ================================
# CHART HELPERS
# ================================

def _chart_colors_js(n: int) -> str:
    colors = CHART_COLORS * (n // len(CHART_COLORS) + 1)
    return json.dumps(colors[:n])


def _render_stacked_bar(month_key: str, shipped: float, forecast: float):
    """水平堆疊長條：已出貨 vs 預計出貨"""
    data = json.dumps({
        "shipped":  round(shipped, 2),
        "forecast": round(forecast, 2),
        "label":    month_key,
    })
    html = f"""
<div style="position:relative;width:100%;height:80px;">
  <canvas id="stackBar"></canvas>
</div>
<div style="display:flex;gap:16px;font-size:12px;color:#888;margin-top:6px;">
  <span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#378ADD;margin-right:4px;"></span>已出貨</span>
  <span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#9FE1CB;margin-right:4px;"></span>預計出貨</span>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
(function(){{
  var d = {data};
  new Chart(document.getElementById('stackBar'), {{
    type: 'bar',
    data: {{
      labels: [d.label],
      datasets: [
        {{ label: '已出貨',   data: [d.shipped],  backgroundColor: '#378ADD', borderRadius: 4 }},
        {{ label: '預計出貨', data: [d.forecast], backgroundColor: '#9FE1CB', borderRadius: 4 }}
      ]
    }},
    options: {{
      indexAxis: 'y', responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: function(c){{ return ' US$' + c.raw.toLocaleString(); }} }} }}
      }},
      scales: {{
        x: {{ stacked: true,
              ticks: {{ callback: function(v){{ return '$' + Math.round(v/1000) + 'k'; }} }},
              grid: {{ color: 'rgba(128,128,128,.12)' }} }},
        y: {{ stacked: true, grid: {{ display: false }} }}
      }}
    }}
  }});
}})();
</script>
"""
    components.html(html, height=110)


def _render_pie_charts(fac_df: pd.DataFrame, cus_df: pd.DataFrame):
    """工廠 / 客戶 甜甜圈圓餅（並排）"""
    fac = fac_df[fac_df["工廠"] != "合計"].copy() if not fac_df.empty else pd.DataFrame()
    cus = cus_df[cus_df["客戶"] != "合計"].copy() if not cus_df.empty else pd.DataFrame()

    fac_labels = json.dumps(fac["工廠"].tolist() if not fac.empty else ["(無資料)"])
    fac_data   = json.dumps([round(v, 2) for v in fac["銷貨金額(USD)"].tolist()] if not fac.empty else [0])
    cus_labels = json.dumps(cus["客戶"].tolist() if not cus.empty else ["(無資料)"])
    cus_data   = json.dumps([round(v, 2) for v in cus["銷貨金額(USD)"].tolist()] if not cus.empty else [0])
    fac_n = max(len(fac), 1)
    cus_n = max(len(cus), 1)
    fac_colors = _chart_colors_js(fac_n)
    cus_colors = _chart_colors_js(cus_n)

    html = f"""
<style>
.pie-wrap {{ display:grid; grid-template-columns:1fr 1fr; gap:20px; }}
.pie-block {{ }}
.pie-title {{ font-size:13px; font-weight:500; color:#555; margin-bottom:6px; }}
.pie-legend {{ display:flex; flex-wrap:wrap; gap:8px; font-size:11px; color:#777; margin-bottom:6px; }}
.pie-dot {{ width:9px; height:9px; border-radius:2px; display:inline-block; margin-right:3px; }}
</style>
<div class="pie-wrap">
  <div class="pie-block">
    <div class="pie-title">依工廠銷貨佔比</div>
    <div class="pie-legend" id="facLeg"></div>
    <div style="position:relative;width:100%;height:200px;"><canvas id="facPie"></canvas></div>
  </div>
  <div class="pie-block">
    <div class="pie-title">依客戶銷貨佔比</div>
    <div class="pie-legend" id="cusLeg"></div>
    <div style="position:relative;width:100%;height:200px;"><canvas id="cusPie"></canvas></div>
  </div>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
(function(){{
  var COLORS = {CHART_COLORS};

  function buildLegend(id, labels, data, colors) {{
    var total = data.reduce(function(a,b){{return a+b;}}, 0);
    var el = document.getElementById(id);
    labels.forEach(function(l,i) {{
      var pct = total > 0 ? (data[i]/total*100).toFixed(1) : '0.0';
      el.innerHTML += '<span style="display:flex;align-items:center;gap:3px">'
        + '<span class="pie-dot" style="background:'+colors[i%colors.length]+'"></span>'
        + l + ' ' + pct + '%</span>';
    }});
  }}

  function makeDoughnut(id, labels, data, colors) {{
    new Chart(document.getElementById(id), {{
      type: 'doughnut',
      data: {{ labels: labels, datasets: [{{
        data: data,
        backgroundColor: colors,
        borderWidth: 1,
        borderColor: 'rgba(255,255,255,.2)'
      }}] }},
      options: {{
        responsive: true, maintainAspectRatio: false, cutout: '58%',
        plugins: {{
          legend: {{ display: false }},
          tooltip: {{ callbacks: {{ label: function(c) {{
            var total = c.dataset.data.reduce(function(a,b){{return a+b;}},0);
            var pct = total > 0 ? (c.raw/total*100).toFixed(1) : '0.0';
            return ' US$' + c.raw.toLocaleString() + ' (' + pct + '%)';
          }} }} }}
        }}
      }}
    }});
  }}

  var facLabels = {fac_labels};
  var facData   = {fac_data};
  var facColors = {fac_colors};
  var cusLabels = {cus_labels};
  var cusData   = {cus_data};
  var cusColors = {cus_colors};

  buildLegend('facLeg', facLabels, facData, facColors);
  buildLegend('cusLeg', cusLabels, cusData, cusColors);
  makeDoughnut('facPie', facLabels, facData, facColors);
  makeDoughnut('cusPie', cusLabels, cusData, cusColors);
}})();
</script>
""".replace("{CHART_COLORS}", json.dumps(CHART_COLORS * 3))
    components.html(html, height=290)


def _render_trend_chart(monthly: list[dict]):
    """近12個月長條+折線組合圖"""
    labels   = json.dumps([r["月份"] for r in monthly])
    shipped  = json.dumps([round(r["已出貨"], 2)   for r in monthly])
    forecast = json.dumps([round(r["預計出貨"], 2)  for r in monthly])
    totals   = json.dumps([round(r["銷貨合計"], 2)  for r in monthly])

    html = f"""
<div style="display:flex;gap:16px;font-size:12px;color:#888;margin-bottom:8px;flex-wrap:wrap;">
  <span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#378ADD;margin-right:4px;"></span>已出貨</span>
  <span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#9FE1CB;margin-right:4px;"></span>預計出貨</span>
  <span><span style="display:inline-block;width:28px;height:2px;background:#D85A30;margin-right:4px;vertical-align:middle;"></span>銷貨合計</span>
</div>
<div style="position:relative;width:100%;height:260px;">
  <canvas id="trendChart"></canvas>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
(function(){{
  new Chart(document.getElementById('trendChart'), {{
    type: 'bar',
    data: {{
      labels: {labels},
      datasets: [
        {{ label: '已出貨',   data: {shipped},  backgroundColor: '#378ADD', order: 2, borderRadius: 3 }},
        {{ label: '預計出貨', data: {forecast}, backgroundColor: '#9FE1CB', order: 2, borderRadius: 3 }},
        {{ label: '銷貨合計', data: {totals},
           type: 'line', borderColor: '#D85A30', backgroundColor: 'transparent',
           pointBackgroundColor: '#D85A30', pointRadius: 4, borderWidth: 2, order: 1 }}
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: function(c){{ return ' US$' + c.raw.toLocaleString(); }} }} }}
      }},
      scales: {{
        x: {{ stacked: true,
              ticks: {{ autoSkip: false, maxRotation: 45, font: {{ size: 11 }} }},
              grid: {{ display: false }} }},
        y: {{ stacked: true,
              ticks: {{ callback: function(v){{ return '$' + Math.round(v/1000) + 'k'; }} }},
              grid: {{ color: 'rgba(128,128,128,.12)' }} }}
      }}
    }}
  }});
}})();
</script>
"""
    components.html(html, height=300)


# ================================
# VIEW SPECS
# ================================

SANDY_NEW_ORDER_SPECS = [
    ("客戶下單日期",          ORDER_DATE_CANDIDATES),
    ("工廠下單日期",          col_candidates("工廠下單日期")),
    ("客戶",                  CUSTOMER_CANDIDATES + ["Customer"]),
    ("PO#",                   PO_CANDIDATES),
    ("P/N",                   PART_CANDIDATES),
    ("Order Q'TY (PCS)",      QTY_CANDIDATES + ["Order QTY (PCS)"]),
    ("Dock",                  col_candidates("Dock")),
    ("Ship date",             PLANNED_SHIP_DATE_CANDIDATES),
    ("WIP",                   WIP_CANDIDATES),
    ("工廠交期",              FACTORY_DUE_CANDIDATES),
    ("交期 (更改)",           col_candidates("交期 (更改)", "交期\n (更改)")),
    ("出貨日期",              ACTUAL_SHIP_DATE_CANDIDATES),
    ("工廠",                  FACTORY_CANDIDATES),
    ("西拓訂單編號",          col_candidates("西拓訂單編號")),
    ("工廠提醒事項",          col_candidates("工廠提醒事項")),
    ("併貨日期 (限內部使用)", col_candidates("併貨日期 (限內部使用)", "併貨日期\n (限內部使用)")),
    ("情況",                  REMARK_CANDIDATES),
    ("客戶要求注意事項",      col_candidates("客戶要求注意事項")),
    ("Ship to",               col_candidates("Ship to")),
    ("Ship via",              col_candidates("Ship via")),
    ("箱數",                  col_candidates("箱數", "CTNS", "CTN")),
    ("重量",                  col_candidates("重量", "Weight", "KGs")),
    ("重貨優惠",              col_candidates("重貨優惠")),
    ("工廠出貨事項",          col_candidates("工廠出貨事項", "工廠出貨注意事項")),
    ("新/舊料號",             col_candidates("新/舊料號")),
    ("板層",                  col_candidates("板層")),
    ("Working Gerber Approval", col_candidates("Working Gerber Approval", "Working\nGerber\nApproval", "WorkingGerberApproval")),
    ("Engineering Question",    col_candidates("Engineering Question", "Engineering\nQuestion", "EngineeringQuestion")),
    ("Pricing & Qty issue",     col_candidates("Pricing & Qty issue")),
    ("T/T",                     col_candidates("T/T")),
]

SANDY_INTERNAL_WIP_SPECS = [
    ("Customer",              CUSTOMER_CANDIDATES + ["Customer"]),
    ("PO#",                   PO_CANDIDATES),
    ("P/N",                   PART_CANDIDATES),
    ("Q'TY (PCS)",            QTY_CANDIDATES + ["Order QTY (PCS)"]),
    ("Dock",                  col_candidates("Dock")),
    ("Ship date",             PLANNED_SHIP_DATE_CANDIDATES),
    ("WIP",                   WIP_CANDIDATES),
    ("出貨狀況 (限內部使用)", col_candidates("出貨狀況 (限內部使用)")),
    ("進度狀況",              col_candidates("進度狀況")),
    ("工廠交期",              FACTORY_DUE_CANDIDATES),
    ("交期 (更改)",           col_candidates("交期 (更改)", "交期\n (更改)")),
    ("出貨日期",              ACTUAL_SHIP_DATE_CANDIDATES),
    ("工廠",                  FACTORY_CANDIDATES),
    ("新/舊料號",             col_candidates("新/舊料號", "新/舊\n料號")),
    ("板層",                  col_candidates("板層", "板\n層")),
    ("工廠出貨事項",          col_candidates("工廠出貨事項", "工廠出貨注意事項")),
    ("西拓訂單編號",          col_candidates("西拓訂單編號")),
    ("工廠提醒事項",          col_candidates("工廠提醒事項")),
    ("併貨日期 (限內部使用)", col_candidates("併貨日期 (限內部使用)")),
    ("客戶要求注意事項",      col_candidates("客戶要求注意事項")),
    ("Ship to",               col_candidates("Ship to")),
    ("Ship via",              col_candidates("Ship via")),
    ("CTN",                     col_candidates("CTN", "CTNS", "箱數")),
    ("KGs",                     col_candidates("KGs", "重量")),
    ("Note",                    REMARK_CANDIDATES),
    ("Working Gerber Approval", col_candidates("Working Gerber Approval", "Working\nGerber\nApproval", "WorkingGerberApproval")),
    ("Engineering Question",    col_candidates("Engineering Question", "Engineering\nQuestion", "EngineeringQuestion")),
    ("Pricing & Qty issue",     col_candidates("Pricing & Qty issue")),
    ("T/T",                     col_candidates("T/T")),
]

SANDY_SALES_SPECS = [
    ("客戶",                  CUSTOMER_CANDIDATES),
    ("PO#",                   PO_CANDIDATES),
    ("P/N",                   PART_CANDIDATES),
    ("Q'TY (PCS)",            QTY_CANDIDATES),
    ("工廠",                  FACTORY_CANDIDATES),
    ("Dock",                  col_candidates("Dock")),
    ("出貨日期",              ACTUAL_SHIP_DATE_CANDIDATES),
    ("Ship date",             PLANNED_SHIP_DATE_CANDIDATES),
    ("工廠交期",              FACTORY_DUE_CANDIDATES),
    ("交期 (更改)",           col_candidates("交期 (更改)", "交期\n (更改)")),
    ("併貨日期 (限內部使用)", col_candidates("併貨日期 (限內部使用)", "併貨日期\n (限內部使用)")),
    ("Ship to",               col_candidates("Ship to")),
    ("Ship via",              col_candidates("Ship via")),
    ("WIP",                   WIP_CANDIDATES),
    ("接單金額",              AMOUNT_ORDER_CANDIDATES),
    ("銷貨金額",              AMOUNT_SHIP_CANDIDATES),
    ("Tracking No.",          col_candidates("Tracking No.", "Tracking No", "TrackingNo.")),
    ("Note",                  REMARK_CANDIDATES),
]


# ================================
# BUILD VIEW DF
# ================================

def build_teable_view_df(source_df: pd.DataFrame, specs):
    mapping = {}
    out = pd.DataFrame(index=source_df.index)
    for display_name, candidates in specs:
        src = find_col(source_df, candidates)
        if src:
            out[display_name] = get_series_by_col(source_df, src)
            mapping[display_name] = src
        else:
            out[display_name] = ""
            mapping[display_name] = None
    out.columns = make_unique_columns(out.columns)
    return out, mapping


# ================================
# CORE FILTER
# ================================

def build_subset_mask(source_df: pd.DataFrame, subset_mode: str) -> pd.Series:
    idx   = source_df.index
    today = pd.Timestamp.today().normalize()

    wip_col   = find_col(source_df, WIP_CANDIDATES)
    wip_raw   = (get_series_by_col(source_df, wip_col).astype(str).str.strip()
                 if wip_col else pd.Series("", index=idx))
    wip_upper = wip_raw.str.upper()

    if subset_mode == "new_order_today":
        cust_col = find_col(source_df, ["客戶下單日期"])
        fact_col = find_col(source_df, ["工廠下單日期"])
        cust_d   = (parse_mixed_date_series(get_series_by_col(source_df, cust_col))
                    if cust_col else pd.Series(pd.NaT, index=idx))
        fact_d   = (parse_mixed_date_series(get_series_by_col(source_df, fact_col))
                    if fact_col else pd.Series(pd.NaT, index=idx))
        return cust_d.dt.normalize().eq(today).fillna(False) | fact_d.dt.normalize().eq(today).fillna(False)

    if subset_mode == "unshipped":
        not_shipment  = ~wip_upper.eq("SHIPMENT")
        not_cancelled = ~_is_cancelled(wip_raw)
        return not_shipment & not_cancelled

    if subset_mode == "shipment_only":
        is_ship    = wip_upper.eq("SHIPMENT")
        actual_col = find_col(source_df, ACTUAL_SHIP_DATE_CANDIDATES)
        actual_d   = (parse_mixed_date_series(get_series_by_col(source_df, actual_col))
                      if actual_col else pd.Series(pd.NaT, index=idx))
        return is_ship & actual_d.dt.normalize().eq(today).fillna(False)

    return pd.Series(True, index=idx)


# ================================
# RENDER TABLE
# ================================

def render_teable_subset_table(title: str, source_df: pd.DataFrame, specs, subset_mode: str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    st.subheader(title)
    mask     = build_subset_mask(source_df, subset_mode)
    filtered = source_df[mask].copy()
    view_df, _ = build_teable_view_df(filtered, specs)
    st.caption(f"共 {len(view_df)} 筆")
    st.dataframe(view_df, use_container_width=True, hide_index=True)

    # ── Excel 下載（自動欄寬）────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.append(list(view_df.columns))
    for row in view_df.itertuples(index=False):
        ws.append(list(row))
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "📥 下載 Excel",
        data=buf,
        file_name=f"{title}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ================================
# 業績明細表（主函式）
# v5.0：加入三個 Tab
# ================================

def render_sales_detail_from_teable(source_df: pd.DataFrame):
    st.subheader("業績明細表")
    if source_df is None or source_df.empty:
        st.info("目前沒有資料。")
        return

    # ── Tab 結構 ─────────────────────────────────────────────────────────────
    if _OBU_AVAILABLE:
        tab1, tab2, tab3 = st.tabs([
            "📊 業績明細",
            "🏷️ 報關單價",
            "🏦 匯 HK OBU 金額",
        ])
    else:
        tab1 = st.container()
        tab2 = tab3 = None

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1：原有業績明細邏輯（完全保留）
    # ══════════════════════════════════════════════════════════════════════════
    with tab1:
        _render_sales_detail_body(source_df)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2：報關單價核算
    # ══════════════════════════════════════════════════════════════════════════
    if _OBU_AVAILABLE and tab2 is not None:
        with tab2:
            render_customs_price_tab()

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3：匯 HK OBU 金額
    # ══════════════════════════════════════════════════════════════════════════
    if _OBU_AVAILABLE and tab3 is not None:
        with tab3:
            render_obu_calc_tab()


# ─────────────────────────────────────────────────────────────────────────────
# 原有業績明細邏輯（抽成獨立函式，邏輯零修改）
# ─────────────────────────────────────────────────────────────────────────────

def _render_sales_detail_body(source_df: pd.DataFrame):
    # ── 欄位偵測 ─────────────────────────────────────────────────────────────
    order_col     = find_col(source_df, ORDER_DATE_CANDIDATES)
    actual_col    = find_col(source_df, ACTUAL_SHIP_DATE_CANDIDATES)
    plan_col      = find_col(source_df, PLANNED_SHIP_DATE_CANDIDATES)
    ship_amt_col  = find_col(source_df, AMOUNT_SHIP_CANDIDATES)
    order_amt_col = find_col(source_df, AMOUNT_ORDER_CANDIDATES)
    customer_col  = find_col(source_df, CUSTOMER_CANDIDATES)
    factory_col   = find_col(source_df, FACTORY_CANDIDATES)
    po_col        = find_col(source_df, PO_CANDIDATES)
    pn_col        = find_col(source_df, PART_CANDIDATES)
    qty_col       = find_col(source_df, QTY_CANDIDATES)
    wip_col       = find_col(source_df, WIP_CANDIDATES)

    # ── 解析 ─────────────────────────────────────────────────────────────────
    order_dates  = parse_mixed_date_series(get_series_by_col(source_df, order_col))
    actual_dates = parse_mixed_date_series(get_series_by_col(source_df, actual_col))
    fallback_col = find_col(source_df, ["出貨日期"])
    if fallback_col and fallback_col != actual_col:
        actual_dates = actual_dates.where(
            actual_dates.notna(),
            parse_mixed_date_series(get_series_by_col(source_df, fallback_col))
        )
    plan_dates   = parse_mixed_date_series(get_series_by_col(source_df, plan_col))
    ship_amount  = parse_amount_series(get_series_by_col(source_df, ship_amt_col))
    order_amount = parse_amount_series(get_series_by_col(source_df, order_amt_col))
    amount       = ship_amount.where(ship_amount.notna() & (ship_amount != 0), order_amount)

    wip       = (get_series_by_col(source_df, wip_col).astype(str).str.upper().str.strip()
                 if wip_col else pd.Series("", index=source_df.index))
    customers = (get_series_by_col(source_df, customer_col).astype(str).fillna("")
                 if customer_col else pd.Series("", index=source_df.index))
    factories = (get_series_by_col(source_df, factory_col).astype(str).fillna("")
                 if factory_col else pd.Series("", index=source_df.index))
    pos       = (get_series_by_col(source_df, po_col).astype(str).fillna("")
                 if po_col else pd.Series("", index=source_df.index))
    pns       = (get_series_by_col(source_df, pn_col).astype(str).fillna("")
                 if pn_col else pd.Series("", index=source_df.index))
    qtys      = (get_series_by_col(source_df, qty_col).astype(str).fillna("")
                 if qty_col else pd.Series("", index=source_df.index))

    # ── 月份選單（近24個月 + 當月，預設當月） ────────────────────────────────
    today          = pd.Timestamp.today().normalize()
    current_period = today.to_period("M")

    valid_periods: set = {current_period}
    for s in [order_dates, actual_dates, plan_dates]:
        for p in s.dt.to_period("M").dropna().unique().tolist():
            if (current_period - p).n <= 24:
                valid_periods.add(p)

    periods = sorted(valid_periods)
    default_index = next(
        (i for i, p in enumerate(periods) if p == current_period),
        len(periods) - 1
    )

    selected = st.selectbox(
        "月份", periods, index=default_index,
        format_func=lambda p: f"{p.year}-{p.month:02d}"
    )
    month_key = f"{selected.year}-{selected.month:02d}"

    # ── Teable 資料過濾 ───────────────────────────────────────────────────────
    is_shipment   = wip.eq("SHIPMENT")
    actual_mask   = actual_dates.dt.to_period("M") == selected
    plan_mask     = plan_dates.dt.to_period("M") == selected
    order_mask    = order_dates.dt.to_period("M") == selected
    shipped_mask  = is_shipment & actual_mask
    forecast_mask = (~is_shipment) & plan_mask

    shipped_df = pd.DataFrame({
        "日期":      actual_dates[shipped_mask].dt.strftime("%Y-%m-%d"),
        "客戶":      customers[shipped_mask].values,
        "工廠":      factories[shipped_mask].values,
        "PO#":       pos[shipped_mask].values,
        "P/N":       pns[shipped_mask].values,
        "QTY":       qtys[shipped_mask].values,
        "WIP":       wip[shipped_mask].values,
        "金額(USD)": amount[shipped_mask].astype(float).values,
    }).reset_index(drop=True)

    forecast_df = pd.DataFrame({
        "日期":      plan_dates[forecast_mask].dt.strftime("%Y-%m-%d"),
        "客戶":      customers[forecast_mask].values,
        "工廠":      factories[forecast_mask].values,
        "PO#":       pos[forecast_mask].values,
        "P/N":       pns[forecast_mask].values,
        "QTY":       qtys[forecast_mask].values,
        "WIP":       wip[forecast_mask].values,
        "金額(USD)": amount[forecast_mask].astype(float).values,
    }).reset_index(drop=True)

    order_total    = float(order_amount[order_mask].fillna(0).sum())
    shipped_total  = float(shipped_df["金額(USD)"].fillna(0).sum())
    forecast_total = float(forecast_df["金額(USD)"].fillna(0).sum())
    legacy_amt     = LEGACY_SHIPPED.get(month_key, 0.0)
    shipped_total += legacy_amt
    month_total    = shipped_total + forecast_total

    # ── 指標卡 ────────────────────────────────────────────────────────────────
    st.markdown(f"### {selected.month}月 業績明細表")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("接單金額 (USD)",     f"${order_total:,.2f}",
              delta=f"{int(order_mask.sum())} 筆")
    c2.metric("已確認出貨 (USD)",   f"${shipped_total:,.2f}",
              delta=f"{len(shipped_df)} 筆 + 歷史 ${legacy_amt:,.2f}" if legacy_amt else f"{len(shipped_df)} 筆")
    c3.metric("預計本月出貨 (USD)", f"${forecast_total:,.2f}",
              delta=f"{len(forecast_df)} 筆")
    c4.metric("月銷貨合計 (USD)",   f"${month_total:,.2f}")

    legacy_note = f"　*（含歷史移轉 US${legacy_amt:,.2f}）*" if legacy_amt else ""
    st.markdown(f"- 已確認出貨（SHIPMENT）：US${shipped_total:,.2f}{legacy_note}")
    st.markdown(f"- 預計{selected.month}月出貨（QA 中）：US${forecast_total:,.2f}")
    st.markdown(f"- {selected.month}月份銷貨金額總計：US${month_total:,.2f}")

    if not forecast_df.empty:
        first = forecast_df.iloc[0]
        try:
            date_str = pd.to_datetime(first["日期"]).strftime("%-m/%-d")
        except Exception:
            date_str = str(first["日期"])
        st.info(
            f"💡 {first['WIP']} 中（預計 {date_str} 出貨）："
            f"{first['客戶']} | US${float(first['金額(USD)']):,.2f}\n\n"
            f"出貨後 WIP 更新為 SHIPMENT，{selected.month}月銷貨合計將增至 US${month_total:,.2f}。"
        )

    # ── ① 已出貨 vs 預計出貨 堆疊長條 ────────────────────────────────────────
    st.markdown("#### 本月出貨進度")
    _render_stacked_bar(month_key, shipped_total, forecast_total)

    # ── 工廠 / 客戶 統計表 ────────────────────────────────────────────────────
    left, right = st.columns(2)

    with left:
        st.markdown(f"#### 🏭 依工廠別統計（{selected.month}月銷貨）")
        if shipped_df.empty:
            msg = "本月 Teable 無已出貨資料。"
            if legacy_amt:
                msg += f"\n歷史移轉金額 US${legacy_amt:,.2f} 已計入合計，無工廠明細。"
            st.info(msg)
            fac_table = pd.DataFrame(columns=["工廠", "訂單數", "銷貨金額(USD)"])
        else:
            fac_table = shipped_df.groupby("工廠", dropna=False).agg(
                訂單數=("PO#", "count"), 銷貨金額=("金額(USD)", "sum")
            ).reset_index()
            fac_table.columns = ["工廠", "訂單數", "銷貨金額(USD)"]
            fac_table = pd.concat([fac_table, pd.DataFrame(
                [["合計", int(fac_table["訂單數"].sum()), float(fac_table["銷貨金額(USD)"].sum())]],
                columns=fac_table.columns)], ignore_index=True)
            st.dataframe(fac_table, use_container_width=True, hide_index=True)

    with right:
        st.markdown(f"#### 👥 依客戶別統計（{selected.month}月銷貨）")
        if shipped_df.empty:
            msg = "本月 Teable 無已出貨資料。"
            if legacy_amt:
                msg += f"\n歷史移轉金額 US${legacy_amt:,.2f} 已計入合計，無客戶明細。"
            st.info(msg)
            cus_table = pd.DataFrame(columns=["客戶", "訂單數", "銷貨金額(USD)"])
        else:
            cus_table = shipped_df.groupby("客戶", dropna=False).agg(
                訂單數=("PO#", "count"), 銷貨金額=("金額(USD)", "sum")
            ).reset_index()
            cus_table.columns = ["客戶", "訂單數", "銷貨金額(USD)"]
            cus_table = pd.concat([cus_table, pd.DataFrame(
                [["合計", int(cus_table["訂單數"].sum()), float(cus_table["銷貨金額(USD)"].sum())]],
                columns=cus_table.columns)], ignore_index=True)
            st.dataframe(cus_table, use_container_width=True, hide_index=True)

    # ── ② 工廠 / 客戶 圓餅圖 ─────────────────────────────────────────────────
    if not shipped_df.empty:
        st.markdown("#### 本月銷貨佔比")
        _render_pie_charts(fac_table, cus_table)

    # ── 明細 ──────────────────────────────────────────────────────────────────
    st.markdown("#### 已出貨明細")
    if shipped_df.empty:
        st.info("無 Teable 出貨明細。")
    else:
        st.dataframe(shipped_df, use_container_width=True, hide_index=True)

    st.markdown("#### 預計出貨明細")
    st.dataframe(forecast_df, use_container_width=True, hide_index=True)

    # ── ③ 近12個月趨勢 ────────────────────────────────────────────────────────
    monthly = []
    for i in range(11, -1, -1):
        p     = current_period - i
        p_key = f"{p.year}-{p.month:02d}"
        s_amt = float(amount[is_shipment & (actual_dates.dt.to_period("M") == p)].fillna(0).sum())
        f_amt = float(amount[(~is_shipment) & (plan_dates.dt.to_period("M") == p)].fillna(0).sum())
        s_amt += LEGACY_SHIPPED.get(p_key, 0.0)
        monthly.append({"月份": p_key, "已出貨": round(s_amt, 2),
                         "預計出貨": round(f_amt, 2), "銷貨合計": round(s_amt + f_amt, 2)})

    st.markdown("#### 近 12 個月月銷貨趨勢")
    _render_trend_chart(monthly)
    st.dataframe(pd.DataFrame(monthly), use_container_width=True, hide_index=True)

    # ── Debug ─────────────────────────────────────────────────────────────────
    with st.expander("Debug：業績明細表欄位偵測"):
        st.json({
            "order_col": order_col, "actual_col": actual_col, "plan_col": plan_col,
            "ship_amt_col": ship_amt_col, "order_amt_col": order_amt_col,
            "customer_col": customer_col, "factory_col": factory_col, "wip_col": wip_col,
            "selected_month": str(selected),
            "teable_shipped": float(amount[is_shipment & actual_mask].fillna(0).sum()),
            "legacy_shipped": legacy_amt,
            "shipped_total": shipped_total,
            "forecast_total": forecast_total,
            "month_total": month_total,
        })


# ================================
# PUBLIC ENTRY POINTS
# ================================

def show_new_orders_wip_report(source_df: pd.DataFrame):
    render_teable_subset_table("📄 新訂單 WIP", source_df, SANDY_NEW_ORDER_SPECS, "new_order_today")


def show_sandy_internal_wip_report(source_df: pd.DataFrame):
    render_teable_subset_table("📄 Sandy 內部 WIP", source_df, SANDY_INTERNAL_WIP_SPECS, "unshipped")


def show_sandy_sales_report(source_df: pd.DataFrame):
    render_teable_subset_table("📄 Sandy 銷貨底", source_df, SANDY_SALES_SPECS, "shipment_only")
